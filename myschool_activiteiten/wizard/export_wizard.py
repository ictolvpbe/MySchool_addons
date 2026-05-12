"""Xlsx-export wizard voor myschool_activiteiten.

Genereert een Excel-overzicht per klas/school met totale kosten,
aantal uitstappen en gemiddelde kost per leerling. Bedoeld voor
boekhouding (budgetopvolging) en directie (overzicht uitstappen).
"""
import base64
import io
from datetime import date

from odoo import api, fields, models
from odoo.exceptions import UserError


def _schoolyear_dates(start_year):
    return date(start_year, 9, 1), date(start_year + 1, 8, 31)


def _current_schoolyear_start():
    today = date.today()
    return today.year if today.month >= 9 else today.year - 1


class ActiviteitenExportWizard(models.TransientModel):
    _name = 'myschool_activiteiten.export.wizard'
    _description = 'Export myschool_activiteiten naar xlsx'

    schoolyear = fields.Integer(
        string='Schooljaar (startjaar)',
        default=lambda self: _current_schoolyear_start(),
        required=True,
        help='Startjaar van het schooljaar. Bv. 2025 = schooljaar 2025-2026.',
    )
    school_id = fields.Many2one(
        'myschool.org', string='School (leeg = alle scholen)',
        domain=[('org_type_id.name', '=', 'SCHOOL')],
    )
    only_done = fields.Boolean(
        string='Enkel afgeronde myschool_activiteiten',
        default=True,
        help='Aangevinkt = enkel state "Afgerond". Uit = ook nog-lopende.',
    )

    def action_export(self):
        self.ensure_one()
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        start, end = _schoolyear_dates(self.schoolyear)
        Act = self.env['myschool_activiteiten.record'].sudo()
        domain = [
            ('datetime', '>=', start),
            ('datetime', '<=', end),
        ]
        if self.only_done:
            domain.append(('state', '=', 'done'))
        if self.school_id:
            domain.append(('school_id', '=', self.school_id.id))
        records = Act.search(domain, order='school_id, datetime')
        if not records:
            raise UserError(
                "Geen myschool_activiteiten gevonden voor schooljaar %d-%d." % (
                    self.schoolyear, self.schoolyear + 1))

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(
            start_color='007d8c', end_color='007d8c', fill_type='solid')
        sub_font = Font(bold=True)
        sub_fill = PatternFill(
            start_color='b0dfe5', end_color='b0dfe5', fill_type='solid')

        # --- Sheet 1: alle myschool_activiteiten in detail ---
        ws = wb.active
        ws.title = f'Detail SJ {self.schoolyear}-{self.schoolyear + 1}'
        headers = [
            'School', 'Datum', 'Referentie', 'Titel', 'Type',
            'Klassen', 'Bestemming',
            'Aantal leerlingen', 'Aanwezig',
            'Vaste kosten', 'Variabele kosten', 'Totale kost',
            'Per aanwezige', 'Per afwezige',
            'S-Code', 'Status',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

        row_idx = 2
        prev_school = None
        section_totals = {'count': 0, 'cost': 0, 'students': 0}
        per_school_summary = {}  # school_name → totals

        for rec in records:
            school_name = rec.school_id.name or '(geen school)'
            if prev_school and school_name != prev_school:
                # Subtotaal-rij vorige school
                self._write_subtotal_row(
                    ws, row_idx, prev_school, section_totals,
                    sub_font, sub_fill)
                per_school_summary[prev_school] = dict(section_totals)
                row_idx += 1
                section_totals = {'count': 0, 'cost': 0, 'students': 0}
            ws.cell(row=row_idx, column=1, value=school_name)
            ws.cell(row=row_idx, column=2,
                    value=rec.datetime.strftime('%d/%m/%Y')
                    if rec.datetime else '')
            ws.cell(row=row_idx, column=3, value=rec.name or '')
            ws.cell(row=row_idx, column=4, value=rec.titel or '')
            ws.cell(row=row_idx, column=5,
                    value=dict(rec._fields['activity_type'].selection).get(
                        rec.activity_type, ''))
            ws.cell(row=row_idx, column=6,
                    value=', '.join(rec.klas_ids.mapped('name')))
            ws.cell(row=row_idx, column=7, value=rec.bestemming_naam or '')
            ws.cell(row=row_idx, column=8, value=rec.aanwezig_total or 0)
            ws.cell(row=row_idx, column=9, value=rec.aanwezig_count or 0)
            ws.cell(row=row_idx, column=10, value=rec.kost_vast_total or 0)
            ws.cell(row=row_idx, column=11, value=rec.kost_variabel_total or 0)
            ws.cell(row=row_idx, column=12, value=rec.totale_kost or 0)
            ws.cell(row=row_idx, column=13, value=rec.kost_per_aanwezig or 0)
            ws.cell(row=row_idx, column=14, value=rec.kost_per_afwezig or 0)
            ws.cell(row=row_idx, column=15, value=rec.s_code_name or '')
            ws.cell(row=row_idx, column=16,
                    value=dict(rec._fields['state'].selection).get(
                        rec.state, ''))
            section_totals['count'] += 1
            section_totals['cost'] += rec.totale_kost or 0
            section_totals['students'] += rec.aanwezig_total or 0
            row_idx += 1
            prev_school = school_name

        if prev_school:
            self._write_subtotal_row(
                ws, row_idx, prev_school, section_totals,
                sub_font, sub_fill)
            per_school_summary[prev_school] = dict(section_totals)
            row_idx += 1

        # Column widths
        widths = [22, 12, 12, 32, 14, 22, 22, 8, 8, 12, 12, 12, 12, 12, 12, 12]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = 'A2'

        # --- Sheet 2: samenvatting per school ---
        ws2 = wb.create_sheet(title='Samenvatting')
        sum_headers = [
            'School', 'Aantal uitstappen', 'Totale kost',
            'Totaal aantal leerlingen', 'Gemiddelde kost per leerling',
        ]
        for col, h in enumerate(sum_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        grand = {'count': 0, 'cost': 0, 'students': 0}
        for r, (school_name, totals) in enumerate(
                sorted(per_school_summary.items()), start=2):
            ws2.cell(row=r, column=1, value=school_name)
            ws2.cell(row=r, column=2, value=totals['count'])
            ws2.cell(row=r, column=3, value=totals['cost'])
            ws2.cell(row=r, column=4, value=totals['students'])
            avg = (totals['cost'] / totals['students']) if totals['students'] else 0
            ws2.cell(row=r, column=5, value=round(avg, 2))
            grand['count'] += totals['count']
            grand['cost'] += totals['cost']
            grand['students'] += totals['students']

        # Grand total
        total_row = len(per_school_summary) + 2
        ws2.cell(row=total_row, column=1, value='TOTAAL').font = (
            Font(bold=True))
        ws2.cell(row=total_row, column=2, value=grand['count']).font = (
            Font(bold=True))
        ws2.cell(row=total_row, column=3, value=grand['cost']).font = (
            Font(bold=True))
        ws2.cell(row=total_row, column=4, value=grand['students']).font = (
            Font(bold=True))
        grand_avg = (grand['cost'] / grand['students']) if grand['students'] else 0
        ws2.cell(row=total_row, column=5, value=round(grand_avg, 2)).font = (
            Font(bold=True))

        for col, w in enumerate([28, 16, 14, 22, 26], 1):
            ws2.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws2.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        file_data = base64.b64encode(output.getvalue())
        output.close()

        filename = (
            f"Activiteiten {self.schoolyear}-{self.schoolyear + 1}.xlsx"
        )
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'mimetype': ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    @staticmethod
    def _write_subtotal_row(ws, row, name, totals, font, fill):
        cell = ws.cell(row=row, column=1, value=f'Subtotaal — {name}')
        cell.font = font
        cell.fill = fill
        for c in range(2, 17):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=8, value=totals['students']).font = font
        ws.cell(row=row, column=12, value=totals['cost']).font = font
