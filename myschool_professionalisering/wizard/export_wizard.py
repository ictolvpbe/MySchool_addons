"""Xlsx-export wizard voor professionaliseringen.

Genereert een Excel-bestand met per leerkracht (en per schooljaar) een
overzicht van opleidingen, uren en status van het bewijs. Bedoeld voor
het jaarverslag van de directie of voor de boekhouding.
"""
import base64
import io
from datetime import date

from odoo import api, fields, models
from odoo.exceptions import UserError


def _schoolyear_dates(start_year):
    """Schooljaar 'start_year' loopt van 1 sept start_year tot 31 aug start_year+1."""
    return date(start_year, 9, 1), date(start_year + 1, 8, 31)


def _current_schoolyear_start():
    today = date.today()
    return today.year if today.month >= 9 else today.year - 1


class ProfessionaliseringExportWizard(models.TransientModel):
    _name = 'myschool_professionalisering.export.wizard'
    _description = 'Export professionaliseringen naar xlsx'

    schoolyear = fields.Integer(
        string='Schooljaar (startjaar)',
        default=lambda self: _current_schoolyear_start(),
        required=True,
        help='Startjaar van het schooljaar. Bv. 2025 = schooljaar 2025-2026 '
             '(1 sept 2025 t.e.m. 31 aug 2026).',
    )
    employee_ids = fields.Many2many(
        'hr.employee', string='Leerkrachten (leeg = allemaal)',
        help='Beperk de export tot specifieke leerkrachten. Leeg = alle '
             'leerkrachten met records in dit schooljaar.',
    )
    only_done = fields.Boolean(
        string='Enkel afgeronde records',
        default=True,
        help='Aangevinkt = enkel records met state "Afgerond" exporteren. '
             'Uit = ook nog-lopende of geweigerde records meenemen.',
    )

    def action_export(self):
        self.ensure_one()
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        start, end = _schoolyear_dates(self.schoolyear)
        Prof = self.env['myschool_professionalisering.record'].sudo()
        domain = [
            ('start_date', '>=', start),
            ('start_date', '<=', end),
        ]
        if self.only_done:
            domain.append(('state', '=', 'done'))
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))
        records = Prof.search(
            domain, order='employee_id, start_date')
        if not records:
            raise UserError(
                "Geen records gevonden voor schooljaar %d-%d." % (
                    self.schoolyear, self.schoolyear + 1))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'SJ {self.schoolyear}-{self.schoolyear + 1}'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(
            start_color='007d8c', end_color='007d8c', fill_type='solid')
        sub_font = Font(bold=True)
        sub_fill = PatternFill(
            start_color='b0dfe5', end_color='b0dfe5', fill_type='solid')

        headers = [
            'Leerkracht', 'Datum', 'Titel', 'Type', 'Vorm', 'Vak',
            'Aanbieder', 'Locatie',
            'Uren', 'Status', 'Bewijs ingediend',
            'Geschatte kost', 'Werkelijke kost',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

        # Group per employee with subtotals
        row_idx = 2
        per_employee_total_uren = 0.0
        prev_employee = None
        section_start_row = 2
        section_total_uren = 0.0
        for rec in records:
            emp_name = rec.employee_id.name or '(geen leerkracht)'
            if prev_employee and emp_name != prev_employee:
                # Subtotaal-rij van vorige medewerker
                self._write_subtotal_row(
                    ws, row_idx, prev_employee, section_total_uren,
                    sub_font, sub_fill)
                row_idx += 1
                section_start_row = row_idx
                section_total_uren = 0.0
            ws.cell(row=row_idx, column=1, value=emp_name)
            ws.cell(row=row_idx, column=2,
                    value=rec.start_date.strftime('%d/%m/%Y')
                    if rec.start_date else '')
            ws.cell(row=row_idx, column=3, value=rec.titel or '')
            ws.cell(row=row_idx, column=4,
                    value=dict(rec._fields['type'].selection).get(
                        rec.type, ''))
            vorm_sel = dict(
                rec._fields['subtype_individueel'].selection)
            ws.cell(row=row_idx, column=5,
                    value=vorm_sel.get(rec.subtype_individueel, ''))
            ws.cell(row=row_idx, column=6,
                    value=rec.vak_id.name if rec.vak_id else '')
            ws.cell(row=row_idx, column=7,
                    value=(rec.address_id.organization
                           if rec.address_id else ''))
            ws.cell(row=row_idx, column=8,
                    value=(rec.address_id.display_address
                           if rec.address_id else ''))
            uren = self._compute_record_uren(rec)
            ws.cell(row=row_idx, column=9, value=uren)
            ws.cell(row=row_idx, column=10,
                    value=dict(rec._fields['state'].selection).get(
                        rec.state, ''))
            ws.cell(row=row_idx, column=11,
                    value='Ja' if rec.bewijs_ingediend else 'Nee')
            ws.cell(row=row_idx, column=12, value=rec.total_cost or 0)
            ws.cell(row=row_idx, column=13, value=rec.s_code_price or 0)
            section_total_uren += uren
            per_employee_total_uren += uren
            row_idx += 1
            prev_employee = emp_name
        # Laatste subtotaal
        if prev_employee:
            self._write_subtotal_row(
                ws, row_idx, prev_employee, section_total_uren,
                sub_font, sub_fill)
            row_idx += 1

        # Grand total
        cell = ws.cell(row=row_idx, column=1, value='TOTAAL')
        cell.font = Font(bold=True, size=12)
        ws.cell(row=row_idx, column=9, value=per_employee_total_uren).font = (
            Font(bold=True, size=12))

        # Column widths
        widths = [28, 12, 36, 12, 14, 18, 22, 28, 8, 14, 10, 14, 14]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        file_data = base64.b64encode(output.getvalue())
        output.close()

        filename = (
            f"Professionalisering {self.schoolyear}-{self.schoolyear + 1}.xlsx"
        )
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'mimetype': ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    @staticmethod
    def _write_subtotal_row(ws, row, name, uren, font, fill):
        cell = ws.cell(row=row, column=1, value=f'Subtotaal — {name}')
        cell.font = font
        cell.fill = fill
        for c in range(2, 14):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=9, value=uren).font = font

    @staticmethod
    def _compute_record_uren(rec):
        """Bereken het aantal uren van een nascholing op basis van
        begin/einduur, of geef een schatting op basis van duur als de
        uren leeg zijn."""
        if rec.start_uur and rec.eind_uur and rec.eind_uur > rec.start_uur:
            return round(rec.eind_uur - rec.start_uur, 2)
        # Fallback op duur-veld
        duur_map = {
            'voormiddag': 3.5,
            'namiddag': 3.5,
            'hele_dag': 7.0,
            'avond': 2.0,
        }
        return duur_map.get(getattr(rec, 'duur', None), 0)
