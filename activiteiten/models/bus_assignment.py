import base64
import io

from odoo import models, fields, api


class ActiviteitenBus(models.Model):
    _name = 'activiteiten.bus'
    _description = 'Busverdeling'
    _order = 'activiteit_id, bus_nummer'

    activiteit_id = fields.Many2one(
        'activiteiten.record', string='Activiteit',
        required=True, ondelete='cascade',
    )
    bus_nummer = fields.Integer(string='Busnummer', required=True, default=1)
    plaatsen = fields.Integer(string='Plaatsen')
    prijs = fields.Monetary(
        string='Prijs', currency_field='currency_id',
        help='Prijs voor deze specifieke bus. Verschillende bussen kunnen '
             'verschillende prijzen hebben (bv. ander type bus, andere route).',
    )
    currency_id = fields.Many2one(
        related='activiteit_id.currency_id', store=False,
    )

    # Available (not yet taken by other buses on same activity)
    beschikbare_klas_ids = fields.Many2many(
        'myschool.org', compute='_compute_beschikbaar', store=False,
        relation='activiteiten_bus_beschikbare_klas_rel',
    )
    beschikbare_leerkracht_ids = fields.Many2many(
        'myschool.person', compute='_compute_beschikbaar', store=False,
        relation='activiteiten_bus_beschikbare_lk_rel',
    )

    # Busbezetting
    klas_ids = fields.Many2many(
        'myschool.org', string='Klassen',
    )
    aantal_klassen = fields.Integer(
        string='Aantal klassen', compute='_compute_counts',
    )
    student_count = fields.Integer(
        string='Aantal leerlingen', compute='_compute_counts',
    )
    aantal_leerkrachten = fields.Integer(
        string='Aantal leerkrachten', compute='_compute_aantal_leerkrachten',
    )
    leerlingen_namen = fields.Text(
        string='Leerlingen', compute='_compute_leerlingen_namen',
    )

    # Verantwoordelijke leerkracht
    busverantwoordelijke_id = fields.Many2one(
        'myschool.person', string='Verantwoordelijke leerkracht',
    )
    busverantwoordelijke_telefoon = fields.Char(
        string='Tel. verantwoordelijke',
        compute='_compute_telefoon', store=True, readonly=False,
    )

    # Leerkracht van de bus = auto-filled with busverantwoordelijke
    leerkracht_bus_id = fields.Many2one(
        'myschool.person', string='Leerkracht van de bus',
        compute='_compute_leerkracht_bus', store=True, readonly=False,
    )

    # Busmaatschappij
    busmaatschappij = fields.Char(string='Busmaatschappij')
    nummerplaat = fields.Char(string='Nummerplaat bus')
    telefoon_chauffeur = fields.Char(string='Tel. buschauffeur')

    # Afwezigen
    afwezigen = fields.Text(string='Afwezigen (naam + klas)')

    @api.depends('activiteit_id.klas_ids', 'activiteit_id.leerkracht_ids',
                 'activiteit_id.bus_ids.klas_ids',
                 'activiteit_id.bus_ids.busverantwoordelijke_id',
                 'activiteit_id.bus_ids.leerkracht_bus_id')
    def _compute_beschikbaar(self):
        for rec in self:
            activiteit = rec.activiteit_id
            other_buses = activiteit.bus_ids - rec
            # Classes already assigned to other buses
            taken_klas = other_buses.mapped('klas_ids')
            rec.beschikbare_klas_ids = activiteit.klas_ids - taken_klas
            # Teachers already assigned to other buses
            taken_lk_ids = (
                other_buses.mapped('busverantwoordelijke_id')
                | other_buses.mapped('leerkracht_bus_id')
            )
            rec.beschikbare_leerkracht_ids = activiteit.leerkracht_ids - taken_lk_ids

    @api.depends('busverantwoordelijke_id')
    def _compute_telefoon(self):
        for rec in self:
            phone = ''
            if rec.busverantwoordelijke_id and rec.busverantwoordelijke_id.odoo_user_id:
                partner = rec.busverantwoordelijke_id.odoo_user_id.partner_id
                phone = partner.phone or ''
            rec.busverantwoordelijke_telefoon = phone

    @api.depends('busverantwoordelijke_id')
    def _compute_leerkracht_bus(self):
        for rec in self:
            rec.leerkracht_bus_id = rec.busverantwoordelijke_id

    @api.depends('klas_ids')
    def _compute_leerlingen_namen(self):
        PropRelation = self.env['myschool.proprelation']
        PropRelationType = self.env['myschool.proprelation.type']
        person_tree_type = PropRelationType.search([('name', '=', 'PERSON-TREE')], limit=1)
        for rec in self:
            if person_tree_type and rec.klas_ids:
                rels = PropRelation.search([
                    ('proprelation_type_id', '=', person_tree_type.id),
                    ('id_org', 'in', rec.klas_ids.ids),
                    ('id_person', '!=', False),
                    ('is_active', '=', True),
                ])
                students = rels.mapped('id_person').sorted('name')
                rec.leerlingen_namen = '\n'.join(students.mapped('name'))
            else:
                rec.leerlingen_namen = ''

    @api.depends('activiteit_id.leerkracht_ids')
    def _compute_aantal_leerkrachten(self):
        for rec in self:
            rec.aantal_leerkrachten = len(rec.activiteit_id.leerkracht_ids)

    @api.depends('klas_ids')
    def _compute_counts(self):
        PropRelation = self.env['myschool.proprelation']
        PropRelationType = self.env['myschool.proprelation.type']
        person_tree_type = PropRelationType.search([('name', '=', 'PERSON-TREE')], limit=1)
        for rec in self:
            rec.aantal_klassen = len(rec.klas_ids)
            if person_tree_type and rec.klas_ids:
                rec.student_count = PropRelation.search_count([
                    ('proprelation_type_id', '=', person_tree_type.id),
                    ('id_org', 'in', rec.klas_ids.ids),
                    ('id_person', '!=', False),
                    ('is_active', '=', True),
                ])
            else:
                rec.student_count = 0

    def write(self, vals):
        res = super().write(vals)
        if 'prijs' in vals:
            # Trigger herberekening van auto-kostenlijnen (bus + verzekering)
            # zodat de Kosten-tab de nieuwe totale busprijs reflecteert.
            self.mapped('activiteit_id')._recalculate_auto_lines()
        return res

    def action_export_busverdeling(self):
        """Export busverdeling as XLSX with a sheet per class."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        records = self.browse(self.env.context.get('active_ids', self.ids))
        if not records:
            return

        activiteit = records[0].activiteit_id
        wb = openpyxl.Workbook()

        # --- Overview sheet ---
        ws = wb.active
        ws.title = 'Overzicht'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='007d8c', end_color='007d8c', fill_type='solid')

        headers = ['Bus', 'Klassen', 'Leerlingen', 'Leerkrachten',
                   'Verantwoordelijke', 'Telefoon', 'Busmaatschappij',
                   'Nummerplaat', 'Tel. chauffeur']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, rec in enumerate(records.sorted('bus_nummer'), 2):
            ws.cell(row=row_idx, column=1, value=rec.bus_nummer)
            ws.cell(row=row_idx, column=2, value=', '.join(rec.klas_ids.mapped('name')))
            ws.cell(row=row_idx, column=3, value=rec.student_count)
            ws.cell(row=row_idx, column=4, value=rec.aantal_leerkrachten)
            ws.cell(row=row_idx, column=5, value=rec.busverantwoordelijke_id.name or '')
            ws.cell(row=row_idx, column=6, value=rec.busverantwoordelijke_telefoon or '')
            ws.cell(row=row_idx, column=7, value=rec.busmaatschappij or '')
            ws.cell(row=row_idx, column=8, value=rec.nummerplaat or '')
            ws.cell(row=row_idx, column=9, value=rec.telefoon_chauffeur or '')

        for col in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

        # --- One sheet per class ---
        PropRelation = self.env['myschool.proprelation']
        PropRelationType = self.env['myschool.proprelation.type']
        person_tree_type = PropRelationType.search([('name', '=', 'PERSON-TREE')], limit=1)

        for rec in records.sorted('bus_nummer'):
            for klas in rec.klas_ids.sorted('name'):
                sheet_name = f"Bus {rec.bus_nummer} - {klas.name}"[:31]
                ws_klas = wb.create_sheet(title=sheet_name)

                # Header row
                klas_headers = ['Nr.', 'Naam', 'Klas']
                for col, h in enumerate(klas_headers, 1):
                    cell = ws_klas.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill

                if person_tree_type:
                    rels = PropRelation.search([
                        ('proprelation_type_id', '=', person_tree_type.id),
                        ('id_org', '=', klas.id),
                        ('id_person', '!=', False),
                        ('is_active', '=', True),
                    ])
                    students = rels.mapped('id_person').sorted('name')
                    for idx, student in enumerate(students, 1):
                        ws_klas.cell(row=idx + 1, column=1, value=idx)
                        ws_klas.cell(row=idx + 1, column=2, value=student.name)
                        ws_klas.cell(row=idx + 1, column=3, value=klas.name)

                ws_klas.column_dimensions['A'].width = 6
                ws_klas.column_dimensions['B'].width = 30
                ws_klas.column_dimensions['C'].width = 12

        # Save to binary
        output = io.BytesIO()
        wb.save(output)
        file_data = base64.b64encode(output.getvalue())
        output.close()

        # Create attachment and return download
        filename = f"Busverdeling - {activiteit.name}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
